# selenium setup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
import html5lib
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time

# navigate to appropriate job listing page
def go_to_listings(driver):

    driver.get("https://jobs.medtronic.com/jobs/search/98857816")
    #https://jobs.medtronic.com/jobs/search/98945694#page2 << page 2 northridge

    #https://jobs.medtronic.com/jobs/search/98857816 << northridge
    #https://jobs.medtronic.com/jobs/search/98830918 << texas
    #//*[@id="notify-modal"]/div[3]/button

    time.sleep(2)

    driver.find_element_by_xpath("//*[@id='notify-modal']/div[3]/button").click()

    time.sleep(5)
 
    driver.find_element_by_xpath("//*[@id='jSearchSubmit']/span[2]").click()

    time.sleep(2)

    driver.execute_script("window.scrollTo(0, 800)") 

    time.sleep(2)

    driver.find_element_by_xpath("//*[@id='job_search_filters']/div[5]/div[2]/div[3]").click()

    time.sleep(2)

# aggregate all url links in page one
def getPageOne(driver):

    time.sleep(5)

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    total_posts = int(soup.find("span", {"class": "total_results"}).text)

    #print(total_posts)

    list_len_int = 10
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    count = 2

    list_post_dates = []

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        list_post_dates.append(post_date_trim)

        count= count +1

    return trimmed_link_list, trimmed_title_list, list_post_dates, total_posts

def getPageTwo(driver):

    time.sleep(5)

    driver.get("https://jobs.medtronic.com/jobs/search/98945694#page2")
    #https://jobs.medtronic.com/jobs/search/98945694#page2 << page 2 northridge

    time.sleep(2)

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    list_len_int = 10
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    count = 2

    list_post_dates = []

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        list_post_dates.append(post_date_trim)

        count= count +1

    return trimmed_link_list, trimmed_title_list, list_post_dates

def getPageThree(driver):

    time.sleep(5)

    driver.get("https://jobs.medtronic.com/jobs/search/98945694#page3")
    #https://jobs.medtronic.com/jobs/search/98945694#page2 << page 2 northridge

    time.sleep(2)

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    list_len_int = 10
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    count = 2

    list_post_dates = []

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        list_post_dates.append(post_date_trim)

        count= count +1

    return trimmed_link_list, trimmed_title_list, list_post_dates

def getPageFour(driver):

    time.sleep(5)

    driver.get("https://jobs.medtronic.com/jobs/search/98945694#page4")
    #https://jobs.medtronic.com/jobs/search/98945694#page2 << page 2 northridge

    time.sleep(2)

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    list_len_int = 10
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    count = 2

    list_post_dates = []

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        list_post_dates.append(post_date_trim)

        count= count +1

    return trimmed_link_list, trimmed_title_list, list_post_dates

def getPageFive(driver, tot_len):

    time.sleep(5)

    driver.get("https://jobs.medtronic.com/jobs/search/98945694#page5")
    #https://jobs.medtronic.com/jobs/search/98945694#page2 << page 2 northridge

    time.sleep(2)

    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    list_len_int = tot_len

    print(list_len_int)
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    count = 2

    list_post_dates = []

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        list_post_dates.append(post_date_trim)

        count= count +1

    return trimmed_link_list, trimmed_title_list, list_post_dates


# 'main' method to iterate through all pages and aggregate URLs
def getURLs():

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(executable_path=r'C:\Users\Michael\Desktop\Automate Application\chromedriver.exe', chrome_options=options)

    go_to_listings(driver)

    time.sleep(5)

    tot_list_title = []
    tot_list_links = []
    tot_list_posts = []

    [page_one_link, page_one_jobs, page_one_post, total_post_num] = getPageOne(driver)

    [page_two_link, page_two_jobs, page_two_post] = getPageTwo(driver)

    [page_three_link, page_three_jobs, page_three_post] = getPageThree(driver)

    [page_four_link, page_four_jobs, page_four_post] = getPageFour(driver)

    last_page_tot = total_post_num - 40

    [page_five_link, page_five_jobs, page_five_post] = getPageFive(driver,last_page_tot)

    tot_list_title = page_one_jobs + page_two_jobs + page_three_jobs + page_four_jobs + page_five_jobs
    tot_list_links = page_one_link + page_two_link + page_three_link + page_four_link+page_five_link
    tot_list_posts = page_one_post + page_two_post + page_three_post+page_four_post + page_five_post

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    ws['A1'] = "Job Title"
    ws['B1'] = "Post Date"
    ws['C1'] = "Job Link"

    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 100

    count = 2

    for i in range(total_post_num):
    
        title_place = 'A' + str(count)
        post_date_place = 'B' + str(count)
        url_place = 'C' + str(count)

        ws[url_place] = str(tot_list_links[i])

        ws[post_date_place] = tot_list_posts[i]

        ws[title_place] = tot_list_title[i]

        print(i)

        count+=1
    
    wb.save('output_Links_medtronic_northridge.xlsx')
    
# for testing purpose
getURLs()
