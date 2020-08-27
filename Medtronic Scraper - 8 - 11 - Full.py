# selenium setup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# open each - grab date, job title, location - these will be different cols in the excel.
# 

from openpyxl import Workbook

import html5lib

from selenium.webdriver.support.ui import Select

# to find links
from bs4 import BeautifulSoup
import urllib.request

import time # to sleep

# navigate to appropriate job listing page
def go_to_listings(driver):

    driver.get("https://jobs.medtronic.com/jobs/search/98830918")
    #https://jobs.medtronic.com/jobs/search/98830918
    #//*[@id="notify-modal"]/div[3]/button

    time.sleep(2)

    driver.find_element_by_xpath("//*[@id='notify-modal']/div[3]/button").click()

    time.sleep(5)
    #driver.find_element_by_class_name('li_location ui-menu-item').click()
 
    driver.find_element_by_xpath("//*[@id='jSearchSubmit']/span[2]").click()

    time.sleep(5)

    driver.find_element_by_xpath("//*[@id='job_search_filters']/div[5]/div[2]/div[3]/a/span[1]").click()

    time.sleep(2)


# aggregate all url links in a set
def aggregate_links(driver):

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    allFixedLinks = [] # all hrefs that exist on the page

    ws['A1'] = "Job Title"
    ws['B1'] = "Post Date"
    ws['C1'] = "Job Link"

    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 100


    time.sleep(5)

    # parse the page source using beautiful soup
    page_source = driver.page_source
    soup = BeautifulSoup(page_source, "html5lib")

    #print(soup.prettify())

    list_len = soup.find("span", {"class": "total_results"}).text

    list_len_int = int(list_len)
    
    # find all hrefs
    allJobLinks = soup.findAll("a", {"class": "job_link font_bold"})

    allJobTitles = [element.text for element in allJobLinks]

    allFixedLinks = [jobLink['href'] for jobLink in allJobLinks]

    #print(allJobLinks)
    #print(len(allJobTitles))

    trimmed_link_list = allFixedLinks[0:list_len_int]
    trimmed_title_list = allJobTitles[0:list_len_int]

    #print(len(trimmed_link_list))

    count = 2
    title_count = 0

    for link in trimmed_link_list:
        
        driver.get(link)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html5lib")

        post_date = soup.find("dd", {"class": "job_post_date"}).text

        post_date_trim = post_date[0:len(post_date) - 10]

        #print(post_date_trim)

        job_title_iter = trimmed_title_list[title_count]

        title_place = 'A' + str(count)
        post_date_place = 'B' + str(count)
        url_place = 'C' + str(count)

        ws[url_place] = str(link)

        ws[post_date_place] = post_date_trim

        ws[title_place] = job_title_iter

        title_count += 1

        count= count +1

    wb.save('output_Links_medtronic.xlsx')



# 'main' method to iterate through all pages and aggregate URLs
def getURLs():

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(executable_path=r'C:\Users\Michael\Desktop\Automate Application\chromedriver.exe', chrome_options=options)

    go_to_listings(driver)

    time.sleep(5)

    aggregate_links(driver)
    
    
# for testing purpose
getURLs()
